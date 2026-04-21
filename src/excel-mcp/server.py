# cspell:words aggfunc fastmcp openpyxl xlsm sheetnames multisheet xlrd xlwt
from __future__ import annotations

import re
from collections.abc import Callable, Sequence
from pathlib import Path
from typing import Any, Literal, TypeVar, cast

import duckdb
import pandas as pd
from mcp.server.fastmcp import FastMCP
from mcp.shared.exceptions import McpError  # ← proper protocol-level errors
from mcp.types import ErrorData, INTERNAL_ERROR
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter, range_boundaries
from openpyxl.utils.cell import column_index_from_string, coordinate_from_string
from openpyxl.worksheet.table import Table, TableStyleInfo

from functools import wraps


mcp = FastMCP("excel-tools")

_F = TypeVar("_F", bound=Callable[..., Any])

# Pandas read_excel / ExcelFile engine parameter (stubs use Literal, not str).
_ExcelReadEngine = Literal["openpyxl", "xlrd"]

_OPENPYXL_SUFFIXES = frozenset({".xlsx", ".xlsm"})
_LEGACY_EXCEL_SUFFIXES = frozenset({".xls"})
_TABULAR_SUFFIXES = frozenset(
    {".xlsx", ".xlsm", ".xls", ".csv", ".parquet"}
)
_EXCEL_SHEET_NAME_MAX = 31

# Maximum rows returned by SQL / filter tools — callers can paginate via OFFSET.
_SQL_ROW_LIMIT = 200


def _is_openpyxl_workbook(path: Path) -> bool:
    return path.suffix.lower() in _OPENPYXL_SUFFIXES


def _is_legacy_excel_path(path: Path) -> bool:
    return path.suffix.lower() in _LEGACY_EXCEL_SUFFIXES


def _excel_read_engine(path: Path) -> _ExcelReadEngine:
    if _is_legacy_excel_path(path):
        return "xlrd"
    if _is_openpyxl_workbook(path):
        return "openpyxl"
    raise ValueError(f"Not an Excel workbook path: {path.suffix}")


def _is_workbook_multiquery_path(path: Path) -> bool:
    """.xlsx / .xlsm / .xls — supported by query_workbook."""
    return _is_openpyxl_workbook(path) or _is_legacy_excel_path(path)


def _legacy_xls_hint(path: Path) -> str:
    if path.suffix.lower() == ".xls":
        return (
            " For .xls files, use convert_legacy_excel_to_xlsx first, "
            "then open the resulting .xlsx."
        )
    return ""


def _keep_vba(path: Path) -> bool:
    return path.suffix.lower() == ".xlsm"


def _sanitize_duckdb_alias(name: str) -> str:
    s = re.sub(r"[^0-9a-zA-Z_]", "_", str(name).strip())
    if not s:
        s = "sheet"
    if s[0].isdigit():
        s = "_" + s
    return s


def _unique_duck_aliases(sheet_names: Sequence[str]) -> dict[str, str]:
    """
    Map original Excel sheet title → unique SQL identifier for DuckDB.

    FIX: The previous implementation could collide when a *natural* alias for
    sheet B happened to match a *generated* suffixed alias for sheet A.
    We now pre-compute all natural aliases first, then resolve collisions in a
    second pass so each sheet gets a deterministic, collision-free name.
    """
    # Pass 1: compute natural (un-suffixed) alias for every sheet.
    natural: list[tuple[str, str]] = [
        (orig, _sanitize_duckdb_alias(orig)) for orig in sheet_names
    ]

    # Pass 2: assign unique names, appending _2, _3, … only when needed.
    used: set[str] = set()
    out: dict[str, str] = {}
    for orig, base in natural:
        candidate = base
        n = 1
        while candidate in used:
            n += 1
            candidate = f"{base}_{n}"
        used.add(candidate)
        out[orig] = candidate
    return out


def _truncate_sheet_title(name: str) -> str:
    n = name[:_EXCEL_SHEET_NAME_MAX]
    if not n.strip():
        raise ValueError("Sheet name cannot be empty")
    return n


def _a1_to_row_col(cell_a1: str) -> tuple[int, int]:
    cell_a1 = cell_a1.replace("$", "").upper().strip()
    col_letter, row = coordinate_from_string(cell_a1)
    return row, column_index_from_string(col_letter)


def _set_cell_value(ws: Any, row: int, column: int, value: Any) -> None:
    """Assign cell value; cast avoids Pylance/MergedCell stub noise."""
    cast(Any, ws.cell(row=row, column=column)).value = value


def _active_sheet(wb: Workbook) -> Any:
    ws = wb.active
    if ws is None:
        raise RuntimeError("Workbook has no active worksheet")
    return ws


def _sanitize_table_display_name(raw: str) -> str:
    """Excel table displayName: letter-first, alphanumeric + underscore."""
    s = re.sub(r"[^0-9a-zA-Z_]", "_", (raw or "Table").strip())
    if not s:
        s = "Table"
    if s[0].isdigit():
        s = "t_" + s
    return s[:255]


def _workbook_table_display_names(wb: Workbook) -> set[str]:
    names: set[str] = set()
    for s in wb.worksheets:
        for t in s.tables.values():
            dn = getattr(t, "displayName", None) or getattr(t, "name", None)
            if dn:
                names.add(str(dn))
    return names


def _allocate_table_display_name(wb: Workbook, base: str) -> str:
    base = _sanitize_table_display_name(base)
    used = _workbook_table_display_names(wb)
    cand = base
    n = 1
    while cand in used:
        n += 1
        suffix = f"_{n}"
        cand = f"{base[: max(1, 255 - len(suffix))]}{suffix}"
    return cand


# ---------------------------------------------------------
# Utility helpers
# ---------------------------------------------------------

def safe_run(fn: _F) -> _F:
    @wraps(fn)
    def wrapper(*args: Any, **kwargs: Any) -> Any:
        try:
            return fn(*args, **kwargs)
        except McpError:
            raise
        except Exception as e:
            raise McpError(ErrorData(code=INTERNAL_ERROR, message=str(e))) from e

    return cast(_F, wrapper)


def load_data(
    file_path: str,
    sheet_name: str | int | None = None,
    *,
    csv_sep: str | None = None,
    csv_encoding: str | None = None,
    csv_decimal: str | None = None,
) -> pd.DataFrame:
    """Load data from Excel (.xlsx / .xlsm / .xls), CSV, or Parquet."""
    path = Path(file_path)

    if not path.exists():
        raise FileNotFoundError(f"{file_path} not found")

    if _is_openpyxl_workbook(path) or _is_legacy_excel_path(path):
        if sheet_name is None:
            sheet_name = 0  # default to first sheet
        engine = _excel_read_engine(path)
        excel_result = pd.read_excel(
            file_path, sheet_name=sheet_name, engine=engine
        )
        if isinstance(excel_result, dict):
            sheets = cast(dict[Any, pd.DataFrame], excel_result)
            if len(sheets) != 1:
                keys = list(sheets.keys())
                preview = keys[:20]
                more = f" (+{len(keys) - 20} more)" if len(keys) > 20 else ""
                raise ValueError(
                    f"Ambiguous multi-sheet read ({len(sheets)} sheets). "
                    f"Pass sheet_name (name or index). Sheets: {preview}{more}"
                )
            df = next(iter(sheets.values()))
        else:
            df = excel_result
        return df

    suf = path.suffix.lower()
    if suf == ".csv":
        csv_kw: dict[str, Any] = {}
        if csv_sep is not None:
            csv_kw["sep"] = csv_sep
        if csv_encoding is not None:
            csv_kw["encoding"] = csv_encoding
        if csv_decimal is not None:
            csv_kw["decimal"] = csv_decimal
        return pd.read_csv(file_path, **csv_kw)

    if suf == ".parquet":
        return pd.read_parquet(file_path)

    raise ValueError(f"Unsupported file type: {path.suffix}")


def validate_columns(df: pd.DataFrame, columns: Sequence[Any]) -> None:
    """Ensure requested columns exist."""
    missing = [c for c in columns if c not in df.columns]
    if missing:
        raise ValueError(f"Columns not found: {missing}")


def _df_to_records(frame: pd.DataFrame | pd.Series) -> list[dict[str, Any]]:
    """Rows as dicts; cast aligns runtime behavior with pandas stubs for orient='records'."""
    tabular: pd.DataFrame = (
        frame if isinstance(frame, pd.DataFrame) else frame.to_frame()
    )
    return cast(list[dict[str, Any]], tabular.to_dict(orient="records"))


def _read_range_as_matrix(
    ws: Any, range_a1: str, max_cells: int
) -> tuple[list[list[Any]], bool, int]:
    """Return (rows as list of lists, truncated, cells_read)."""
    min_col, min_row, max_col, max_row = cast(
        tuple[int, int, int, int],
        range_boundaries(range_a1.strip()),
    )
    rows_out: list[list[Any]] = []
    count = 0
    for r in range(min_row, max_row + 1):
        row_vals: list[Any] = []
        for c in range(min_col, max_col + 1):
            if count >= max_cells:
                return rows_out, True, count
            row_vals.append(ws.cell(row=r, column=c).value)
            count += 1
        rows_out.append(row_vals)
    return rows_out, False, count


def _matrix_to_records(matrix: list[list[Any]]) -> list[dict[str, Any]]:
    if not matrix:
        return []
    headers: list[str] = []
    for i, h in enumerate(matrix[0]):
        if h is None or (isinstance(h, float) and pd.isna(h)):
            headers.append(f"column_{i + 1}")
        else:
            headers.append(str(h))
    records: list[dict[str, Any]] = []
    for row in matrix[1:]:
        rec: dict[str, Any] = {}
        for i, name in enumerate(headers):
            rec[name] = row[i] if i < len(row) else None
        records.append(rec)
    return records


def _safe_write_sheet(
    path: Path,
    sheet_name: str,
    df: pd.DataFrame,
    *,
    excel_table: bool = False,
    table_display_name: str | None = None,
    table_style: str = "TableStyleMedium2",
) -> None:
    """
    Write tabular data to one sheet using openpyxl (headers + rows).

    The target sheet is removed and re-created, so anything that lived only on
    that sheet (charts, tables, formatting on that tab) is lost. Other
    worksheets are untouched, which avoids pandas ExcelWriter re-saving the
    whole book in a way that can disturb workbook-level metadata.

    When excel_table is True and the frame has at least one data row, an Excel
    Table (ListObject) is added over the header + body range with banded rows.
    """
    st = _truncate_sheet_title(sheet_name)

    if not path.exists():
        wb = Workbook()
        ws = _active_sheet(wb)
        ws.title = st
    else:
        wb = load_workbook(path, keep_vba=_keep_vba(path))
        if st in wb.sheetnames:
            # Delete and recreate to clear stale data while keeping other sheets.
            del wb[st]
        ws = wb.create_sheet(st)

    # Write header row.
    for col_idx, col_name in enumerate(df.columns, start=1):
        _set_cell_value(ws, 1, col_idx, col_name)

    # Write data rows.
    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        for col_idx, val in enumerate(row, start=1):
            # Convert numpy scalars to plain Python so openpyxl is happy.
            if hasattr(val, "item"):
                val = val.item()  # type: ignore[assignment]
            _set_cell_value(ws, row_idx, col_idx, val)

    if excel_table and not df.empty and len(df.columns) > 0:
        nrows = len(df) + 1
        ncols = len(df.columns)
        ref = f"A1:{get_column_letter(ncols)}{nrows}"
        base_nm = table_display_name or f"tbl_{sheet_name}"
        disp = _allocate_table_display_name(wb, base_nm)
        tab = Table(displayName=disp, ref=ref)
        tab.tableStyleInfo = TableStyleInfo(
            name=table_style,
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        ws.add_table(tab)

    wb.save(path)
    wb.close()


# ---------------------------------------------------------
# Dataset discovery
# ---------------------------------------------------------


@mcp.tool()
@safe_run
def list_datasets(folder_path: str) -> list[str]:
    """List datasets in a folder."""
    p = Path(folder_path)
    if not p.is_dir():
        raise NotADirectoryError(f"Not a directory or does not exist: {folder_path}")

    return [
        str(f)
        for f in p.glob("*")
        if f.is_file() and f.suffix.lower() in _TABULAR_SUFFIXES
    ]


# ---------------------------------------------------------
# Basic dataset inspection
# ---------------------------------------------------------


@mcp.tool()
@safe_run
def inspect_excel(file_path: str) -> dict[str, Any]:
    """Return sheet names, columns, types, and preview rows."""
    path = Path(file_path)
    if not _is_openpyxl_workbook(path):
        raise ValueError(
            f"inspect_excel only supports .xlsx and .xlsm workbooks.{_legacy_xls_hint(path)}"
        )

    with pd.ExcelFile(file_path, engine="openpyxl") as xls:
        info: dict[str, Any] = {}

        for sheet in xls.sheet_names:
            df = pd.read_excel(
                file_path, sheet_name=sheet, nrows=20, engine="openpyxl"
            )

            info[str(sheet)] = {
                "columns": list(df.columns),
                "dtypes": df.dtypes.astype(str).to_dict(),
                "preview_rows": _df_to_records(df.head(5)),
            }

        return info


@mcp.tool()
@safe_run
def inspect_csv(
    file_path: str,
    csv_sep: str | None = None,
    csv_encoding: str | None = None,
    csv_decimal: str | None = None,
) -> dict[str, Any]:
    """Return columns, dtypes, and preview rows for a CSV (reads first 20 rows)."""
    path = Path(file_path)
    if path.suffix.lower() != ".csv":
        raise ValueError("inspect_csv only supports .csv files")
    if not path.exists():
        raise FileNotFoundError(file_path)

    csv_kw: dict[str, Any] = {"nrows": 20}
    if csv_sep is not None:
        csv_kw["sep"] = csv_sep
    if csv_encoding is not None:
        csv_kw["encoding"] = csv_encoding
    if csv_decimal is not None:
        csv_kw["decimal"] = csv_decimal
    df = pd.read_csv(file_path, **csv_kw)

    return {
        "columns": list(df.columns),
        "dtypes": df.dtypes.astype(str).to_dict(),
        "preview_rows": _df_to_records(df.head(5)),
    }


@mcp.tool()
@safe_run
def workbook_structure(file_path: str) -> dict[str, Any]:
    """
    List sheets, used dimensions, and Excel Table (ListObject) definitions.
    For .xlsx / .xlsm only (openpyxl).
    """
    path = Path(file_path)
    if not _is_openpyxl_workbook(path):
        raise ValueError(
            f"workbook_structure requires .xlsx or .xlsm.{_legacy_xls_hint(path)}"
        )

    wb = load_workbook(
        file_path, data_only=True, keep_vba=_keep_vba(path), read_only=False
    )
    try:
        sheets: dict[str, Any] = {}
        for sn in wb.sheetnames:
            ws = wb[sn]
            tables = [
                {
                    "name": t.name,
                    "display_name": t.displayName,
                    "ref": t.ref,
                }
                for t in ws.tables.values()
            ]
            sheets[sn] = {
                "max_row": ws.max_row,
                "max_column": ws.max_column,
                "tables": tables,
            }
        return {"sheets": sheets}
    finally:
        wb.close()


@mcp.tool()
@safe_run
def read_excel_range(
    file_path: str,
    sheet_name: str,
    range_a1: str,
    value_mode: str = "computed",
    max_cells: int = 20_000,
    offset_row: int = 0,
) -> dict[str, Any]:
    """
    Read a rectangular A1-style range as a 2D grid, with optional row offset
    for pagination.

    value_mode:
      "computed" — evaluated cell values (data_only=True, uses Excel's cached
                   results). NOTE: cached values are only present when the file
                   was last saved by Excel itself. Files written by openpyxl or
                   other non-Excel tools will return None for formula cells in
                   this mode. Use "stored" to retrieve the raw formula text.
      "stored"   — cell contents as saved (formula text like '=SUM(A1:A2)'
                   where applicable).

    offset_row: skip this many data rows before returning (0-based). Use
    together with max_cells to paginate large ranges.
    """
    path = Path(file_path)
    if not _is_openpyxl_workbook(path):
        raise ValueError(
            f"read_excel_range requires .xlsx or .xlsm.{_legacy_xls_hint(path)}"
        )
    if value_mode not in ("computed", "stored"):
        raise ValueError('value_mode must be "computed" or "stored"')

    data_only = value_mode == "computed"
    wb = load_workbook(
        file_path,
        data_only=data_only,
        keep_vba=_keep_vba(path),
        read_only=False,
    )
    try:
        if sheet_name not in wb.sheetnames:
            raise ValueError(
                f"Sheet {sheet_name!r} not found. Available: {wb.sheetnames}"
            )
        ws = wb[sheet_name]
        matrix, truncated, cells = _read_range_as_matrix(
            ws, range_a1, max_cells=max_cells
        )

        # Apply row offset for pagination (keep header row intact).
        if offset_row and len(matrix) > 1:
            matrix = [matrix[0]] + matrix[1 + offset_row :]

        return {
            "range": range_a1.strip(),
            "value_mode": value_mode,
            "data_only_caveat": (
                "Cached values only. Cells computed by non-Excel tools may be None."
                if value_mode == "computed"
                else None
            ),
            "truncated": truncated,
            "cells_read": cells,
            "offset_row": offset_row,
            "row_count": len(matrix),
            "column_count": len(matrix[0]) if matrix else 0,
            "rows": matrix,
        }
    finally:
        wb.close()


@mcp.tool()
@safe_run
def read_excel_table(
    file_path: str,
    sheet_name: str,
    table_name: str,
    value_mode: str = "computed",
    max_cells: int = 100_000,
) -> dict[str, Any]:
    """
    Read an Excel Table (ListObject) by name or display name on a sheet.
    Returns header-based row records plus table metadata.

    NOTE: When value_mode is "computed", cached formula results are used.
    These are only populated when the file was last saved by Excel. See
    read_excel_range for full details on this caveat.
    """
    path = Path(file_path)
    if not _is_openpyxl_workbook(path):
        raise ValueError(
            f"read_excel_table requires .xlsx or .xlsm.{_legacy_xls_hint(path)}"
        )
    if value_mode not in ("computed", "stored"):
        raise ValueError('value_mode must be "computed" or "stored"')

    data_only = value_mode == "computed"
    wb = load_workbook(
        file_path,
        data_only=data_only,
        keep_vba=_keep_vba(path),
        read_only=False,
    )
    try:
        if sheet_name not in wb.sheetnames:
            raise ValueError(
                f"Sheet {sheet_name!r} not found. Available: {wb.sheetnames}"
            )
        ws = wb[sheet_name]
        ref: str | None = None
        resolved_name: str | None = None
        for t in ws.tables.values():
            if t.name == table_name or t.displayName == table_name:
                ref = t.ref
                resolved_name = t.displayName or t.name
                break
        if ref is None:
            names = [t.displayName or t.name for t in ws.tables.values()]
            raise ValueError(
                f"Table {table_name!r} not on sheet {sheet_name!r}. "
                f"Known tables: {names}"
            )
        matrix, truncated, cells = _read_range_as_matrix(
            ws, ref, max_cells=max_cells
        )
        records = _matrix_to_records(matrix)
        return {
            "table": resolved_name,
            "ref": ref,
            "value_mode": value_mode,
            "truncated": truncated,
            "cells_read": cells,
            "row_count": len(records),
            "rows": records,
        }
    finally:
        wb.close()


@mcp.tool()
@safe_run
def dataset_summary(
    file_path: str,
    sheet_name: str | int | None = None,
    csv_sep: str | None = None,
    csv_encoding: str | None = None,
    csv_decimal: str | None = None,
) -> dict[str, Any]:
    """Return row count, column names, dtypes, and per-column null counts."""
    df = load_data(
        file_path,
        sheet_name,
        csv_sep=csv_sep,
        csv_encoding=csv_encoding,
        csv_decimal=csv_decimal,
    )

    return {
        "rows": len(df),
        "columns": list(df.columns),
        "dtypes": df.dtypes.astype(str).to_dict(),
        "missing_values": df.isnull().sum().to_dict(),
    }


@mcp.tool()
@safe_run
def dataset_shape(
    file_path: str,
    sheet_name: str | int | None = None,
    csv_sep: str | None = None,
    csv_encoding: str | None = None,
    csv_decimal: str | None = None,
) -> dict[str, int]:
    """Return dataset shape."""
    df = load_data(
        file_path,
        sheet_name,
        csv_sep=csv_sep,
        csv_encoding=csv_encoding,
        csv_decimal=csv_decimal,
    )

    return {
        "rows": len(df),
        "columns": len(df.columns),
    }


@mcp.tool()
@safe_run
def sample_rows(
    file_path: str,
    sheet_name: str | int | None = None,
    n: int = 20,
    csv_sep: str | None = None,
    csv_encoding: str | None = None,
    csv_decimal: str | None = None,
) -> list[dict[str, Any]]:
    """Return random sample rows."""
    df = load_data(
        file_path,
        sheet_name,
        csv_sep=csv_sep,
        csv_encoding=csv_encoding,
        csv_decimal=csv_decimal,
    )

    if len(df) == 0:
        return []

    return _df_to_records(df.sample(min(n, len(df))))


# ---------------------------------------------------------
# Reading tools
# ---------------------------------------------------------


@mcp.tool()
@safe_run
def read_sheet(
    file_path: str,
    sheet_name: str | int | None = None,
    csv_sep: str | None = None,
    csv_encoding: str | None = None,
    csv_decimal: str | None = None,
) -> dict[str, Any]:
    """
    Read a sheet (Excel) or full file (CSV / Parquet). For Excel, sheet_name
    defaults to the first sheet when omitted. For CSV / Parquet, sheet_name is ignored.
    """
    df = load_data(
        file_path,
        sheet_name,
        csv_sep=csv_sep,
        csv_encoding=csv_encoding,
        csv_decimal=csv_decimal,
    )

    return {
        "rows": len(df),
        "columns": list(df.columns),
        "sample": _df_to_records(df.head(10)),
    }


@mcp.tool()
@safe_run
def summarize_dataset(
    file_path: str,
    sheet_name: str | int | None = None,
    csv_sep: str | None = None,
    csv_encoding: str | None = None,
    csv_decimal: str | None = None,
) -> dict[str, Any]:
    """Return pandas descriptive statistics (describe), including non-numeric columns."""
    df = load_data(
        file_path,
        sheet_name,
        csv_sep=csv_sep,
        csv_encoding=csv_encoding,
        csv_decimal=csv_decimal,
    )

    return df.describe(include="all").to_dict()


# ---------------------------------------------------------
# Data filtering
# ---------------------------------------------------------


@mcp.tool()
@safe_run
def filter_rows(
    file_path: str,
    column: str,
    value: Any,
    *,
    sheet_name: str | int | None = None,
    operator: str = "==",
    limit: int = 100,
    offset: int = 0,
    csv_sep: str | None = None,
    csv_encoding: str | None = None,
    csv_decimal: str | None = None,
) -> list[dict[str, Any]]:
    """
    Filter rows using a column condition with optional pagination.

    operators: ==, >, <, >=, <=, !=, contains
    limit:  max rows to return (default 100).
    offset: skip this many matching rows before returning (for pagination).
    For Excel, pass sheet_name (name or index) or omit for the first sheet.
    For CSV / Parquet, sheet_name is ignored.
    """
    df = load_data(
        file_path,
        sheet_name,
        csv_sep=csv_sep,
        csv_encoding=csv_encoding,
        csv_decimal=csv_decimal,
    )
    validate_columns(df, [column])

    coerced: Any = value
    if operator in ("==", "!=", ">", "<", ">=", "<="):
        try:
            coerced = float(value)
        except (ValueError, TypeError):
            coerced = value

    ops: dict[str, Any] = {
        "==": lambda col: df[col] == coerced,
        "!=": lambda col: df[col] != coerced,
        ">":  lambda col: df[col] > coerced,
        "<":  lambda col: df[col] < coerced,
        ">=": lambda col: df[col] >= coerced,
        "<=": lambda col: df[col] <= coerced,
        "contains": lambda col: df[col].astype(str).str.contains(
            str(value), regex=False
        ),
    }

    if operator not in ops:
        raise ValueError(
            f"Unsupported operator {operator!r}. "
            f"Valid: {sorted(ops)}"
        )

    filtered = df[ops[operator](column)]
    page = filtered.iloc[offset : offset + limit]

    return _df_to_records(page)


# ---------------------------------------------------------
# Pivot tables
# ---------------------------------------------------------


@mcp.tool()
@safe_run
def pivot_table(
    file_path: str,
    index: str | list[str],
    values: str | list[str],
    *,
    sheet_name: str | int | None = None,
    agg: str = "sum",
    csv_sep: str | None = None,
    csv_encoding: str | None = None,
    csv_decimal: str | None = None,
) -> list[dict[str, Any]]:
    """Create a pivot table. For Excel, omit sheet_name for the first sheet."""
    df = load_data(
        file_path,
        sheet_name,
        csv_sep=csv_sep,
        csv_encoding=csv_encoding,
        csv_decimal=csv_decimal,
    )

    index_cols = [index] if isinstance(index, str) else list(index)
    value_cols = [values] if isinstance(values, str) else list(values)

    validate_columns(df, index_cols + value_cols)

    pivot = cast(
        pd.DataFrame,
        pd.pivot_table(
            df,
            index=index_cols,
            values=value_cols,
            aggfunc=cast(Any, agg),
        ),
    )

    return _df_to_records(pivot.reset_index())


# ---------------------------------------------------------
# Column analysis
# ---------------------------------------------------------


@mcp.tool()
@safe_run
def top_values(
    file_path: str,
    column: str,
    sheet_name: str | int | None = None,
    n: int = 10,
    csv_sep: str | None = None,
    csv_encoding: str | None = None,
    csv_decimal: str | None = None,
) -> dict[Any, int]:
    """Return most common values in a column."""
    df = load_data(
        file_path,
        sheet_name,
        csv_sep=csv_sep,
        csv_encoding=csv_encoding,
        csv_decimal=csv_decimal,
    )
    validate_columns(df, [column])

    counts = df[column].value_counts().head(n)

    return counts.to_dict()


# ---------------------------------------------------------
# SQL querying (LIMIT/OFFSET applied by the tool around your SELECT)
# ---------------------------------------------------------


@mcp.tool()
@safe_run
def query_dataset(
    file_path: str,
    query: str,
    sheet_name: str | int | None = None,
    limit: int = _SQL_ROW_LIMIT,
    offset: int = 0,
    csv_sep: str | None = None,
    csv_encoding: str | None = None,
    csv_decimal: str | None = None,
) -> dict[str, Any]:
    """
    Run SQL on the loaded table registered as "data".

    Loads one sheet from .xlsx / .xlsm / .xls (pass sheet_name as name or index;
    default: first sheet), or a CSV / Parquet file (sheet_name ignored).
    For joins across Excel sheets use query_workbook.

    Pagination: use limit and offset to page through large result sets.
    The tool appends LIMIT/OFFSET automatically — do NOT add them to your query.

    ⚠ SQL runs locally in DuckDB against an in-memory DataFrame. Do not pass
    untrusted user input in production environments.
    """
    df = load_data(
        file_path,
        sheet_name,
        csv_sep=csv_sep,
        csv_encoding=csv_encoding,
        csv_decimal=csv_decimal,
    )

    paginated_query = f"SELECT * FROM ({query}) _q LIMIT {limit} OFFSET {offset}"

    con = duckdb.connect()
    try:
        con.register("data", df)
        result = con.execute(paginated_query).df()
    finally:
        con.close()

    return {
        "limit": limit,
        "offset": offset,
        "returned_rows": len(result),
        "rows": _df_to_records(result),
    }


@mcp.tool()
@safe_run
def query_workbook(
    file_path: str,
    query: str,
    sheet_names: list[str] | None = None,
    limit: int = _SQL_ROW_LIMIT,
    offset: int = 0,
) -> dict[str, Any]:
    """
    Load multiple Excel sheets into DuckDB and run one SQL query.
    Supports .xlsx, .xlsm, and legacy .xls (via xlrd).

    Each sheet is registered under a SQL-safe name (see registered_tables).
    Sanitized names replace non-alphanumeric characters with underscores;
    collisions get numeric suffixes (_2, _3, …). Quote identifiers in SQL if
    needed. If sheet_names is omitted, all sheets in the workbook are loaded.

    Pagination: use limit and offset. Do NOT add LIMIT/OFFSET to your query.

    FIX: alias collision resolution now correctly handles cases where a
    natural alias for one sheet matches a suffixed alias for another.
    """
    path = Path(file_path)
    if not _is_workbook_multiquery_path(path):
        raise ValueError(
            "query_workbook supports .xlsx, .xlsm, and .xls workbooks only"
        )

    engine = _excel_read_engine(path)
    with pd.ExcelFile(file_path, engine=engine) as xlf:
        all_names: list[str] = [str(x) for x in list(xlf.sheet_names)]
        if sheet_names is not None:
            chosen: list[str] = [str(s) for s in sheet_names]
        else:
            chosen = list(all_names)
        for sn in chosen:
            if sn not in all_names:
                raise ValueError(f"Unknown sheet {sn!r}. Available: {all_names}")

        aliases = _unique_duck_aliases(chosen)
        paginated_query = (
            f"SELECT * FROM ({query}) _q LIMIT {limit} OFFSET {offset}"
        )
        con = duckdb.connect()
        try:
            for orig in chosen:
                df = pd.read_excel(
                    file_path, sheet_name=orig, engine=engine
                )
                con.register(aliases[orig], df)
            result = con.execute(paginated_query).df()
        finally:
            con.close()

        return {
            "registered_tables": aliases,
            "limit": limit,
            "offset": offset,
            "returned_rows": len(result),
            "rows": _df_to_records(result),
        }


@mcp.tool()
@safe_run
def convert_legacy_excel_to_xlsx(
    source_path: str,
    output_path: str,
    if_file_exists: str = "replace",
) -> str:
    """
    Convert a legacy .xls workbook to .xlsx (all sheets) using pandas + openpyxl.
    Sheet names are truncated to 31 characters; collisions after truncation raise.
    """
    src = Path(source_path)
    out = Path(output_path)
    if not _is_legacy_excel_path(src):
        raise ValueError("source_path must be a .xls file")
    if out.suffix.lower() != ".xlsx":
        raise ValueError("output_path must end with .xlsx")
    if not src.exists():
        raise FileNotFoundError(source_path)
    if out.exists() and if_file_exists == "error":
        raise FileExistsError(output_path)

    xls_engine: Literal["xlrd"] = "xlrd"
    with pd.ExcelFile(source_path, engine=xls_engine) as xlf:
        raw_names: list[str] = [str(x) for x in list(xlf.sheet_names)]

    out.parent.mkdir(parents=True, exist_ok=True)
    seen: set[str] = set()
    with pd.ExcelWriter(out, engine="openpyxl", mode="w") as writer:
        for raw in raw_names:
            sn = _truncate_sheet_title(raw)
            if sn in seen:
                raise ValueError(
                    f"Duplicate sheet name after truncation: {sn!r}"
                )
            seen.add(sn)
            df = pd.read_excel(
                source_path, sheet_name=raw, engine=xls_engine
            )
            df.to_excel(writer, sheet_name=sn, index=False)

    return (
        f"Converted {source_path!r} → {output_path!r} ({len(raw_names)} sheet(s))."
    )


# ---------------------------------------------------------
# Writing tools
# ---------------------------------------------------------


@mcp.tool()
@safe_run
def write_sheet(
    file_path: str,
    sheet_name: str,
    data: list[dict[str, Any]],
    as_excel_table: bool = False,
    table_display_name: str | None = None,
    table_style: str = "TableStyleMedium2",
) -> str:
    """
    Write rows to a sheet; creates a new workbook if the file does not exist.

    Uses openpyxl via _safe_write_sheet: other sheets are left alone; the
    named sheet is replaced (see _safe_write_sheet for what is preserved).

    Set as_excel_table=True to format the range as an Excel Table (ListObject)
    with header row and banded rows (table_style defaults to TableStyleMedium2).
    """
    path = Path(file_path)
    path.parent.mkdir(parents=True, exist_ok=True)
    st = _truncate_sheet_title(sheet_name)
    df = pd.DataFrame(data)
    _safe_write_sheet(
        path,
        st,
        df,
        excel_table=as_excel_table,
        table_display_name=table_display_name,
        table_style=table_style,
    )
    extra = " as an Excel Table" if as_excel_table and not df.empty else ""
    return f"Sheet '{st}' written ({len(data)} rows){extra}."


@mcp.tool()
@safe_run
def append_rows_to_sheet(
    file_path: str,
    sheet_name: str,
    data: list[dict[str, Any]],
) -> str:
    """
    Append rows to a sheet using openpyxl (preserves other sheets and VBA
    containers in .xlsm better than full replace). Row 1 must be headers
    when the sheet already has data; new sheets get headers from dict keys.
    """
    if not data:
        return "No rows to append."
    path = Path(file_path)
    if not _is_openpyxl_workbook(path):
        raise ValueError(
            f"append_rows_to_sheet requires .xlsx or .xlsm.{_legacy_xls_hint(path)}"
        )
    path.parent.mkdir(parents=True, exist_ok=True)
    st = _truncate_sheet_title(sheet_name)
    keys = list(data[0].keys())

    if not path.exists():
        wb = Workbook()
        ws = _active_sheet(wb)
        ws.title = st
        for j, k in enumerate(keys, start=1):
            _set_cell_value(ws, 1, j, k)
        for ri, row in enumerate(data, start=2):
            for j, k in enumerate(keys, start=1):
                _set_cell_value(ws, ri, j, row.get(k))
        wb.save(file_path)
        wb.close()
        return f"Created {file_path!r} with sheet {st!r} ({len(data)} rows)."

    wb = load_workbook(file_path, keep_vba=_keep_vba(path))
    try:
        if st not in wb.sheetnames:
            ws = wb.create_sheet(st)
            for j, k in enumerate(keys, start=1):
                _set_cell_value(ws, 1, j, k)
            for ri, row in enumerate(data, start=2):
                for j, k in enumerate(keys, start=1):
                    _set_cell_value(ws, ri, j, row.get(k))
            wb.save(file_path)
            return f"Added sheet {st!r} with {len(data)} rows."

        ws = wb[st]
        any_val = any(
            any(v is not None for v in row)
            for row in ws.iter_rows(
                min_row=1, max_row=ws.max_row, values_only=True
            )
        )

        if not any_val:
            for j, k in enumerate(keys, start=1):
                _set_cell_value(ws, 1, j, k)
            for ri, row in enumerate(data, start=2):
                for j, k in enumerate(keys, start=1):
                    _set_cell_value(ws, ri, j, row.get(k))
            wb.save(file_path)
            return f"Sheet {st!r} was empty; wrote header and {len(data)} rows."

        col_by: dict[str, int] = {}
        for c in range(1, ws.max_column + 1):
            h = cast(Any, ws.cell(row=1, column=c)).value
            if h is not None:
                label = str(h)
                if label not in col_by:
                    col_by[label] = c
        if not col_by:
            raise ValueError("Could not read headers from row 1.")
        missing = [str(k) for k in keys if str(k) not in col_by]
        if missing:
            raise ValueError(
                f"Keys not in sheet header: {missing}. Known: {sorted(col_by)}"
            )
        next_r = ws.max_row + 1
        for i, row in enumerate(data):
            r = next_r + i
            for k in keys:
                _set_cell_value(ws, r, col_by[str(k)], row.get(k))
        wb.save(file_path)
        return f"Appended {len(data)} rows to {st!r}."
    finally:
        wb.close()


@mcp.tool()
@safe_run
def write_multisheet_workbook(
    output_path: str,
    sheets: dict[str, list[dict[str, Any]]],
    if_file_exists: str = "replace",
) -> str:
    """
    Create or overwrite a workbook with multiple sheets from row dicts.
    Sheet names are truncated to 31 characters (Excel limit). if_file_exists:
    'replace' overwrites; 'error' raises if the file exists.
    """
    out = Path(output_path)
    if not _is_openpyxl_workbook(out):
        raise ValueError(
            f"output_path must end with .xlsx or .xlsm.{_legacy_xls_hint(out)}"
        )
    if out.exists() and if_file_exists == "error":
        raise FileExistsError(output_path)
    if not sheets:
        raise ValueError("sheets must contain at least one sheet")

    out.parent.mkdir(parents=True, exist_ok=True)
    seen: set[str] = set()
    with pd.ExcelWriter(out, engine="openpyxl", mode="w") as writer:
        for raw_name, rows in sheets.items():
            sn = _truncate_sheet_title(raw_name)
            if sn in seen:
                raise ValueError(
                    f"Duplicate sheet name after truncation: {sn!r}"
                )
            seen.add(sn)
            pd.DataFrame(rows).to_excel(writer, sheet_name=sn, index=False)
    return f"Wrote {len(sheets)} sheet(s) to {output_path}."


@mcp.tool()
@safe_run
def write_csv(
    file_path: str,
    data: list[dict[str, Any]],
    sep: str = ",",
    encoding: str = "utf-8",
    index: bool = False,
) -> str:
    """Write row dicts to a CSV file (overwrites)."""
    path = Path(file_path)
    if path.suffix.lower() != ".csv":
        raise ValueError("file_path must end with .csv")
    path.parent.mkdir(parents=True, exist_ok=True)
    pd.DataFrame(data).to_csv(file_path, sep=sep, encoding=encoding, index=index)
    return f"Wrote {len(data)} row(s) to {file_path!r}."


@mcp.tool()
@safe_run
def write_range_matrix(
    file_path: str,
    sheet_name: str,
    start_cell: str,
    data: list[list[Any]],
) -> str:
    """
    Write a 2D list into the sheet starting at start_cell (e.g. B2). Values
    that are strings beginning with '=' are written as Excel formulas.
    Creates the file or sheet if needed.
    """
    path = Path(file_path)
    if not _is_openpyxl_workbook(path):
        raise ValueError(
            f"write_range_matrix requires .xlsx or .xlsm.{_legacy_xls_hint(path)}"
        )
    path.parent.mkdir(parents=True, exist_ok=True)
    st = _truncate_sheet_title(sheet_name)
    start_row, start_col = _a1_to_row_col(start_cell)

    if not path.exists():
        wb = Workbook()
        ws = _active_sheet(wb)
        ws.title = st
    else:
        wb = load_workbook(path, keep_vba=_keep_vba(path))
        if st not in wb.sheetnames:
            wb.create_sheet(st)
        ws = wb[st]

    try:
        for ri, row in enumerate(data):
            for ci, val in enumerate(row):
                _set_cell_value(ws, start_row + ri, start_col + ci, val)
        wb.save(file_path)
    finally:
        wb.close()

    ncols = max((len(r) for r in data), default=0)
    return (
        f"Wrote {len(data)}x{ncols} matrix at {start_cell!r} on sheet {st!r}."
    )


@mcp.tool()
@safe_run
def add_workbook_sheet(file_path: str, sheet_name: str) -> str:
    """Add an empty worksheet to an existing .xlsx / .xlsm file."""
    path = Path(file_path)
    if not path.exists():
        raise FileNotFoundError(file_path)
    if not _is_openpyxl_workbook(path):
        raise ValueError(
            f"add_workbook_sheet requires .xlsx or .xlsm.{_legacy_xls_hint(path)}"
        )
    st = _truncate_sheet_title(sheet_name)
    wb = load_workbook(path, keep_vba=_keep_vba(path))
    try:
        if st in wb.sheetnames:
            raise ValueError(f"Sheet {st!r} already exists")
        wb.create_sheet(st)
        wb.save(file_path)
    finally:
        wb.close()
    return f"Added empty sheet {st!r}."


@mcp.tool()
@safe_run
def delete_sheet(file_path: str, sheet_name: str) -> str:
    """
    NEW — Delete a worksheet from an existing .xlsx / .xlsm workbook.
    Raises if the sheet does not exist or if it is the only sheet.
    """
    path = Path(file_path)
    if not path.exists():
        raise FileNotFoundError(file_path)
    if not _is_openpyxl_workbook(path):
        raise ValueError(
            f"delete_sheet requires .xlsx or .xlsm.{_legacy_xls_hint(path)}"
        )
    wb = load_workbook(path, keep_vba=_keep_vba(path))
    try:
        if sheet_name not in wb.sheetnames:
            raise ValueError(
                f"Sheet {sheet_name!r} not found. Available: {wb.sheetnames}"
            )
        if len(wb.sheetnames) == 1:
            raise ValueError("Cannot delete the only sheet in a workbook.")
        del wb[sheet_name]
        wb.save(file_path)
    finally:
        wb.close()
    return f"Deleted sheet {sheet_name!r}."


@mcp.tool()
@safe_run
def rename_sheet(file_path: str, old_name: str, new_name: str) -> str:
    """
    NEW — Rename a worksheet in an existing .xlsx / .xlsm workbook.
    New name is truncated to 31 characters per Excel's limit.
    """
    path = Path(file_path)
    if not path.exists():
        raise FileNotFoundError(file_path)
    if not _is_openpyxl_workbook(path):
        raise ValueError(
            f"rename_sheet requires .xlsx or .xlsm.{_legacy_xls_hint(path)}"
        )
    st_new = _truncate_sheet_title(new_name)
    wb = load_workbook(path, keep_vba=_keep_vba(path))
    try:
        if old_name not in wb.sheetnames:
            raise ValueError(
                f"Sheet {old_name!r} not found. Available: {wb.sheetnames}"
            )
        if st_new in wb.sheetnames and st_new != old_name:
            raise ValueError(f"Sheet {st_new!r} already exists.")
        wb[old_name].title = st_new
        wb.save(file_path)
    finally:
        wb.close()
    return f"Renamed {old_name!r} → {st_new!r}."


@mcp.tool()
@safe_run
def create_summary_sheet(
    file_path: str,
    output_file: str,
    group_by_column: str | None = None,
    sheet_name: str | int | None = None,
    csv_sep: str | None = None,
    csv_encoding: str | None = None,
    csv_decimal: str | None = None,
) -> str:
    """Create a workbook with raw data and a numeric sum grouped by one column."""
    df = load_data(
        file_path,
        sheet_name,
        csv_sep=csv_sep,
        csv_encoding=csv_encoding,
        csv_decimal=csv_decimal,
    )

    if df.shape[1] == 0:
        raise ValueError("Dataset has no columns to group by")
    key = group_by_column if group_by_column is not None else df.columns[0]
    validate_columns(df, [key])

    summary = cast(
        pd.DataFrame,
        df.groupby(key, dropna=False).sum(numeric_only=True),
    )

    out = Path(output_file)
    out.parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="data", index=False)
        summary.to_excel(writer, sheet_name="summary")

    return f"Created {output_file}"


# ---------------------------------------------------------
# Start MCP server
# ---------------------------------------------------------

if __name__ == "__main__":
    mcp.run(transport="stdio")
